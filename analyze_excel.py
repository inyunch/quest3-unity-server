import openpyxl
import sys

# Read Excel file
wb = openpyxl.load_workbook(r'C:\Users\user\temp_log.xlsx')
ws = wb.active

# Print header
header = [cell.value for cell in ws[1]]
print("Column headers:", header)
print("\n" + "="*150 + "\n")

# Print first 10 data rows
print(f"{'Row':<5} {'upload_ms':<12} {'download_ms':<12} {'upload_bytes':<15} {'download_bytes':<15} {'upload_comp':<15} {'download_comp':<15}")
print("="*150)

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=11, values_only=True), start=2):
    # Find column indices
    upload_ms_idx = header.index('upload_ms') if 'upload_ms' in header else None
    download_ms_idx = header.index('download_ms') if 'download_ms' in header else None

    # Try to find the byte columns (they might have different names)
    upload_bytes_idx = None
    download_bytes_idx = None
    upload_bytes_uncomp_idx = None
    download_bytes_uncomp_idx = None
    upload_bytes_comp_idx = None
    download_bytes_comp_idx = None

    for idx, col in enumerate(header):
        if col and 'upload_bytes' in col.lower():
            if 'uncompressed' in col.lower():
                upload_bytes_uncomp_idx = idx
            elif 'compressed' in col.lower():
                upload_bytes_comp_idx = idx
            else:
                upload_bytes_idx = idx
        elif col and 'download_bytes' in col.lower():
            if 'uncompressed' in col.lower():
                download_bytes_uncomp_idx = idx
            elif 'compressed' in col.lower():
                download_bytes_comp_idx = idx
            else:
                download_bytes_idx = idx

    upload_ms = row[upload_ms_idx] if upload_ms_idx is not None else 'N/A'
    download_ms = row[download_ms_idx] if download_ms_idx is not None else 'N/A'

    # Get upload bytes (prefer uncompressed, fallback to regular)
    upload_bytes = row[upload_bytes_uncomp_idx] if upload_bytes_uncomp_idx is not None else (row[upload_bytes_idx] if upload_bytes_idx is not None else 'N/A')

    # Get download bytes (prefer uncompressed, fallback to regular)
    download_bytes = row[download_bytes_uncomp_idx] if download_bytes_uncomp_idx is not None else (row[download_bytes_idx] if download_bytes_idx is not None else 'N/A')

    # Get compressed sizes
    upload_comp = row[upload_bytes_comp_idx] if upload_bytes_comp_idx is not None else 'N/A'
    download_comp = row[download_bytes_comp_idx] if download_bytes_comp_idx is not None else 'N/A'

    print(f"{i:<5} {str(upload_ms):<12} {str(download_ms):<12} {str(upload_bytes):<15} {str(download_bytes):<15} {str(upload_comp):<15} {str(download_comp):<15}")

print("\n" + "="*150)
print("\nColumn mapping found:")
print(f"  upload_ms: column {upload_ms_idx}")
print(f"  download_ms: column {download_ms_idx}")
print(f"  upload_bytes (uncompressed): column {upload_bytes_uncomp_idx}")
print(f"  upload_bytes_compressed: column {upload_bytes_comp_idx}")
print(f"  download_bytes (uncompressed): column {download_bytes_uncomp_idx}")
print(f"  download_bytes_compressed: column {download_bytes_comp_idx}")
